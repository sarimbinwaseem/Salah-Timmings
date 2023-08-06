from openpyxl import load_workbook
import shelve
import datetime

# Making months list: date given \/ is just to make legal date.
months = [datetime.datetime(2023, i, 3).strftime("%B") for i in range(1, 13)]

# Opening every month's file and dumping data to shelve object. 
for month in months:

	workbook = load_workbook(f'../XLSX/{month}.xlsx')
	sheet = workbook.active

	date = 1
	with shelve.open(f"../Times/{month}") as db:

		for row in sheet.iter_rows(values_only = True):

			# row is a tuple.
			db[str(date)] = row
			date += 1
